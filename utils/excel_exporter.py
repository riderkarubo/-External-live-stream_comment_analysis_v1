"""Excelファイル出力モジュール"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
from io import BytesIO
import pandas as pd
from config import (
    COLOR_MAP,
    CHAT_ATTRIBUTES,
    CHAT_SENTIMENTS,
    ANSWER_STATUSES
)


def rgb_to_hex(rgb_dict: Dict[str, float]) -> str:
    """
    RGB辞書を16進数カラーコードに変換
    
    Args:
        rgb_dict: {"red": 0.0-1.0, "green": 0.0-1.0, "blue": 0.0-1.0}
        
    Returns:
        16進数カラーコード（例: "FF0000"）
    """
    r = int(rgb_dict.get("red", 0.5) * 255)
    g = int(rgb_dict.get("green", 0.5) * 255)
    b = int(rgb_dict.get("blue", 0.5) * 255)
    return f"{r:02X}{g:02X}{b:02X}"


def create_excel_file(
    df: pd.DataFrame,
    question_df: pd.DataFrame,
    statistics: Dict,
    question_statistics: Dict,
    output_path: str,
    progress_callback: Optional[callable] = None
) -> str:
    """
    Excelファイルを作成
    
    Args:
        df: 全コメントのデータフレーム
        question_df: 質問コメントのデータフレーム
        statistics: 統計情報
        question_statistics: 質問統計情報
        output_path: 出力ファイルパス
        progress_callback: 進捗コールバック関数
        
    Returns:
        出力ファイルパス
    """
    wb = Workbook()
    
    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # メインシートを作成
    if progress_callback:
        progress_callback("メインシートを作成中...", 0.1)
    
    create_main_excel_sheet(wb, df, statistics, progress_callback)
    
    # 質問シートを作成（質問がある場合のみ）
    if len(question_df) > 0:
        if progress_callback:
            progress_callback("質問シートを作成中...", 0.7)
        
        create_question_excel_sheet(wb, question_df, question_statistics, progress_callback)
    
    # ファイルを保存
    if progress_callback:
        progress_callback("Excelファイルを保存中...", 0.9)
    
    wb.save(output_path)
    
    if progress_callback:
        progress_callback("Excelファイルの作成が完了しました", 1.0)
    
    return output_path


def create_excel_file_in_memory(
    df: pd.DataFrame,
    question_df: pd.DataFrame,
    statistics: Dict,
    question_statistics: Dict,
    progress_callback: Optional[callable] = None
) -> bytes:
    """
    Excelファイルをメモリ上で作成（BytesIOを使用）
    
    Args:
        df: 全コメントのデータフレーム
        question_df: 質問コメントのデータフレーム
        statistics: 統計情報
        question_statistics: 質問統計情報
        progress_callback: 進捗コールバック関数
        
    Returns:
        Excelファイルのバイトデータ
    """
    wb = Workbook()
    
    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # メインシートを作成
    create_main_excel_sheet(wb, df, statistics, progress_callback)
    
    # 質問シートを作成（質問がある場合のみ）
    if len(question_df) > 0:
        create_question_excel_sheet(wb, question_df, question_statistics, progress_callback)
    
    # メモリ上に保存
    if progress_callback:
        progress_callback("Excelファイルを保存中...", 0.9)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_data = output.getvalue()
    output.close()
    
    if progress_callback:
        progress_callback("Excelファイルの作成が完了しました", 1.0)
    
    return file_data


def create_main_excel_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    statistics: Dict,
    progress_callback: Optional[callable] = None
):
    """
    メインシートをExcelに作成
    
    Args:
        wb: Workbookオブジェクト
        df: データフレーム
        statistics: 統計情報
        progress_callback: 進捗コールバック関数
    """
    ws = wb.create_sheet("メインシート", 0)
    
    # 統計情報を書き込み
    row = 1
    ws.cell(row=row, column=1, value="=== 統計情報 ===")
    row += 1
    ws.cell(row=row, column=1, value=f"全コメント件数: {statistics['total_comments']}")
    row += 1
    row += 1  # 空行
    
    ws.cell(row=row, column=1, value="【チャットの属性別件数】")
    row += 1
    for attr, count in statistics["attribute_counts"].items():
        ws.cell(row=row, column=1, value=f"{attr}: {count}件")
        row += 1
    
    row += 1  # 空行
    ws.cell(row=row, column=1, value="【チャット感情別件数】")
    row += 1
    for sentiment, count in statistics["sentiment_counts"].items():
        ws.cell(row=row, column=1, value=f"{sentiment}: {count}件")
        row += 1
    
    row += 1  # 空行
    row += 1  # 空行
    ws.cell(row=row, column=1, value="=== コメントデータ ===")
    row += 1
    
    # ヘッダー行
    headers = ["guest_id", "username", "original_text", "inserted_at", "チャットの属性", "チャット感情"]
    header_row = row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
    row += 1
    
    # データ行を書き込み
    data_start_row = row
    
    if progress_callback:
        progress_callback("データを書き込んでいます...", 0.3)
    
    for idx, (_, data_row) in enumerate(df.iterrows()):
        ws.cell(row=row, column=1, value=str(data_row.get("guest_id", "")))
        ws.cell(row=row, column=2, value=str(data_row.get("username", "")))
        ws.cell(row=row, column=3, value=str(data_row.get("original_text", "")))
        ws.cell(row=row, column=4, value=str(data_row.get("inserted_at", "")))
        ws.cell(row=row, column=5, value=str(data_row.get("チャットの属性", "")))
        ws.cell(row=row, column=6, value=str(data_row.get("チャット感情", "")))
        row += 1
        
        # 進捗更新（1000行ごと）
        if (idx + 1) % 1000 == 0 and progress_callback:
            progress = 0.3 + (idx + 1) / len(df) * 0.3
            progress_callback(f"データを書き込んでいます... ({idx + 1}/{len(df)})", progress)
    
    data_end_row = row - 1
    
    if progress_callback:
        progress_callback("書式設定を適用中...", 0.6)
    
    # ドロップダウンと色分けを適用
    # username列（列B）
    username_col = 2
    apply_dropdown(ws, username_col, data_start_row, data_end_row, df["username"].unique().tolist())
    
    # チャットの属性列（列E）
    attribute_col = 5
    apply_dropdown(ws, attribute_col, data_start_row, data_end_row, CHAT_ATTRIBUTES)
    apply_color_formatting_excel(
        ws, attribute_col, data_start_row, data_end_row,
        {attr: COLOR_MAP.get(attr, {}) for attr in CHAT_ATTRIBUTES},
        df, "チャットの属性"
    )
    
    # チャット感情列（列F）
    sentiment_col = 6
    apply_dropdown(ws, sentiment_col, data_start_row, data_end_row, CHAT_SENTIMENTS)
    apply_color_formatting_excel(
        ws, sentiment_col, data_start_row, data_end_row,
        {sent: COLOR_MAP.get(sent, {}) for sent in CHAT_SENTIMENTS},
        df, "チャット感情"
    )
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20


def create_question_excel_sheet(
    wb: Workbook,
    question_df: pd.DataFrame,
    question_statistics: Dict,
    progress_callback: Optional[callable] = None
):
    """
    質問シートをExcelに作成
    
    Args:
        wb: Workbookオブジェクト
        question_df: 質問コメントのデータフレーム
        question_statistics: 質問統計情報
        progress_callback: 進捗コールバック関数
    """
    ws = wb.create_sheet("質問シート")
    
    # 統計情報を書き込み
    row = 1
    ws.cell(row=row, column=1, value="=== 統計情報 ===")
    row += 1
    ws.cell(row=row, column=1, value=f"質問コメント件数: {question_statistics['total_questions']}")
    row += 1
    ws.cell(row=row, column=1, value=f"質問回答率: {question_statistics['answer_rate']:.1f}%")
    row += 1
    row += 1  # 空行
    row += 1  # 空行
    ws.cell(row=row, column=1, value="=== 質問コメントデータ ===")
    row += 1
    
    # ヘッダー行
    headers = ["guest_id", "username", "original_text", "inserted_at", "チャットの属性", "チャット感情", "回答状況"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
    row += 1
    
    # データ行を書き込み
    data_start_row = row
    
    for _, data_row in question_df.iterrows():
        ws.cell(row=row, column=1, value=str(data_row.get("guest_id", "")))
        ws.cell(row=row, column=2, value=str(data_row.get("username", "")))
        ws.cell(row=row, column=3, value=str(data_row.get("original_text", "")))
        ws.cell(row=row, column=4, value=str(data_row.get("inserted_at", "")))
        ws.cell(row=row, column=5, value=str(data_row.get("チャットの属性", "")))
        ws.cell(row=row, column=6, value=str(data_row.get("チャット感情", "")))
        ws.cell(row=row, column=7, value=str(data_row.get("回答状況", "未回答")))
        row += 1
    
    data_end_row = row - 1
    
    if progress_callback:
        progress_callback("書式設定を適用中...", 0.85)
    
    # 回答状況列にドロップダウンと色分けを適用（列G）
    answer_status_col = 7
    apply_dropdown(ws, answer_status_col, data_start_row, data_end_row, ANSWER_STATUSES)
    apply_color_formatting_excel(
        ws, answer_status_col, data_start_row, data_end_row,
        {status: COLOR_MAP.get(status, {}) for status in ANSWER_STATUSES},
        question_df, "回答状況"
    )
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15


def apply_dropdown(ws, column: int, start_row: int, end_row: int, options: List[str]):
    """
    ドロップダウンを適用
    
    Args:
        ws: Worksheetオブジェクト
        column: 列番号（1始まり）
        start_row: 開始行
        end_row: 終了行
        options: オプションリスト
    """
    if end_row < start_row:
        return
    
    col_letter = get_column_letter(column)
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"
    
    # openpyxlのDataValidationでは、カンマ区切りのリストとして指定
    # オプションをエスケープしてカンマで結合
    escaped_options = [opt.replace('"', '""') for opt in options]
    options_str = ",".join(escaped_options)
    dv = DataValidation(type="list", formula1=f'"{options_str}"', allow_blank=True)
    dv.error = "選択された値はリストにありません"
    dv.errorTitle = "無効な入力"
    dv.prompt = "リストから選択してください"
    dv.promptTitle = "選択"
    
    ws.add_data_validation(dv)
    dv.add(range_str)


def apply_color_formatting_excel(
    ws,
    column: int,
    start_row: int,
    end_row: int,
    value_color_map: Dict[str, Dict],
    df: pd.DataFrame,
    column_name: str
):
    """
    Excelシートに色付けを適用
    
    Args:
        ws: Worksheetオブジェクト
        column: 列番号（1始まり）
        start_row: 開始行
        end_row: 終了行
        value_color_map: 値と色のマッピング
        df: データフレーム
        column_name: 列名
    """
    if end_row < start_row:
        return
    
    col_letter = get_column_letter(column)
    
    # 各行に色を適用
    for idx, (_, row_data) in enumerate(df.iterrows()):
        row_num = start_row + idx
        cell_value = str(row_data.get(column_name, "")).strip()
        
        # 値に対応する色を取得
        color = None
        for key in value_color_map.keys():
            if key in cell_value or cell_value == key:
                color = value_color_map[key]
                break
        
        if color:
            hex_color = rgb_to_hex(color)
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell = ws[f"{col_letter}{row_num}"]
            cell.fill = fill

